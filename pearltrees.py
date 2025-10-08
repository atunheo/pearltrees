# streamlit_app.py
import streamlit as st
import requests
from bs4 import BeautifulSoup
import pandas as pd
import openpyxl
from openpyxl import Workbook
import os

OUTPUT_FILE = "pearltrees_urls.xlsx"

def fetch_pearltrees_links(username):
    """Lấy danh sách URL bài viết từ trang Pearltrees"""
    base_url = f"https://www.pearltrees.com/{username}"
    response = requests.get(base_url)
    if response.status_code != 200:
        st.error("❌ Không thể truy cập trang Pearltrees. Vui lòng kiểm tra tên người dùng.")
        return []

    soup = BeautifulSoup(response.text, "html.parser")
    links = []

    # Pearltrees hiển thị liên kết trong thẻ <a> — ta lọc URL phù hợp
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.startswith("/"):
            href = f"https://www.pearltrees.com{href}"
        if "n=" in href or "itemId" in href or "id=" in href:
            links.append(href)

    return list(set(links))  # loại trùng

def save_urls_to_excel(urls, filename=OUTPUT_FILE):
    """Lưu danh sách URL vào file Excel"""
    if not urls:
        return

    if not os.path.exists(filename):
        wb = Workbook()
        ws = wb.active
        ws.title = "URLs"
        ws.append(["Index", "URL"])
    else:
        wb = openpyxl.load_workbook(filename)
        ws = wb.active

    existing = {ws.cell(row=i, column=2).value for i in range(2, ws.max_row + 1)}
    new_urls = [u for u in urls if u not in existing]

    for url in new_urls:
        ws.append([ws.max_row, url])

    wb.save(filename)
    st.success(f"✅ Đã lưu {len(new_urls)} URL mới vào file Excel.")

def load_urls(filename=OUTPUT_FILE):
    """Đọc danh sách URL"""
    if os.path.exists(filename):
        return pd.read_excel(filename)
    return pd.DataFrame(columns=["Index", "URL"])

# --- Giao diện Streamlit ---
st.title("🌐 Trình thu thập URL từ Pearltrees")
username = st.text_input("Nhập tên người dùng Pearltrees:", placeholder="vd: heiliaounu")

if st.button("🔍 Thu thập link bài viết"):
    if not username.strip():
        st.warning("⚠️ Vui lòng nhập tên người dùng.")
    else:
        with st.spinner("Đang thu thập dữ liệu..."):
            urls = fetch_pearltrees_links(username.strip())
            if urls:
                st.success(f"✅ Tìm thấy {len(urls)} URL.")
                save_urls_to_excel(urls)
                st.dataframe(pd.DataFrame(urls, columns=["URL"]))
            else:
                st.info("Không tìm thấy liên kết nào trên trang này.")

st.subheader("📄 Danh sách URL đã lưu")
urls_df = load_urls()
st.dataframe(urls_df)

if os.path.exists(OUTPUT_FILE):
    with open(OUTPUT_FILE, "rb") as f:
        st.download_button(
            label="📥 Tải xuống file Excel",
            data=f.read(),
            file_name=OUTPUT_FILE,
        )
